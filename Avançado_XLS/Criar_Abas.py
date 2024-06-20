from openpyxl import load_workbook
from openpyxl import Workbook as WB
import string
import os

wb_dados = load_workbook(filename='Quebrar.xlsx')
sheetDados = wb_dados['Dados']

nomeCliente = ""
totalLinha = len(sheetDados['A']) + 1

for linha in range(2, totalLinha):
    nomeAtual = sheetDados['A%s' % linha].value

    if nomeAtual == nomeCliente:
        linhaSheetQuebra = len(sheetCliente['A']) + 1

        cColumA = "A" + str(linhaSheetQuebra)
        cColumB = "B" + str(linhaSheetQuebra)
        cColumC = "C" + str(linhaSheetQuebra)
        
        sheetCliente[cColumA] = sheetDados['A%s' % linhaSheetQuebra].value
        sheetCliente[cColumB] = sheetDados['B%s' % linhaSheetQuebra].value
        sheetCliente[cColumC] = sheetDados['C%s' % linhaSheetQuebra].value
        
    else:

        wb_dados.create_sheet(title=nomeAtual)
        sheetCliente = wb_dados[nomeAtual]

        nomeCliente = sheetDados['A%s' % linha].value

        for letra in string.ascii_uppercase:
            coluna = letra+'1'

            if sheetDados[coluna] != "":
                sheetCliente[coluna] = sheetDados[coluna].value
            else:
                break
        
        sheetCliente['A2'] = sheetDados['A%s' % linha].value
        sheetCliente['B2'] = sheetDados['B%s' % linha].value
        sheetCliente['C2'] = sheetDados['C%s' % linha].value


wb_dados.save(filename='Quebrar.xlsx')
os.startfile('Quebrar.xlsx')