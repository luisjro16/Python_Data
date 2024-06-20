from openpyxl import load_workbook
from openpyxl import Workbook as WB

from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles import colors
from openpyxl.cell import Cell

import string
import os

wb_dados = load_workbook(filename='Quebrar.xlsx')
sheetDados = wb_dados['Dados']

nomeCliente = ""

corAzul = PatternFill(start_color='0099CCFF',
                      end_color='0099CCFF',
                      fill_type='solid')

corAmarelo = PatternFill(start_color='00FFFFCC',
                      end_color='00FFFFCC',
                      fill_type='solid')

borda_fina = Side(border_style="thin",  color="000000")
borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)

for linha in range(2, len(sheetDados['A']) + 1):
    nomeAtual = sheetDados['A%s' % linha].value

    if nomeAtual == nomeCliente:
        linhaSheetQuebra = len(sheetRelatorio['A']) + 1

        cColumA = "A" + str(linhaSheetQuebra)
        cColumB = "B" + str(linhaSheetQuebra)
        cColumC = "C" + str(linhaSheetQuebra)

        sheetRelatorio[cColumA] = sheetDados['A%s' % linhaSheetQuebra].value
        sheetRelatorio[cColumB] = sheetDados['B%s' % linhaSheetQuebra].value
        sheetRelatorio[cColumC] = sheetDados['C%s' % linhaSheetQuebra].value

        sheetRelatorio[cColumA].fill = corAmarelo
        sheetRelatorio[cColumB].fill = corAmarelo
        sheetRelatorio[cColumC].fill = corAmarelo

        sheetRelatorio[cColumA].border = borda
        sheetRelatorio[cColumB].border = borda
        sheetRelatorio[cColumC].border = borda

        wb_cliente.save(filename=nomeCliente+'.xlsx')
    else:

        wb_cliente = WB()
        sheetRelatorio = wb_cliente.active
        sheetRelatorio.title = 'Relatorio'

        nomeCliente = sheetDados['A%s' % linha].value

        for letra in string.ascii_uppercase:
            coluna = letra+'1'

            if sheetDados[coluna] != "":
                sheetRelatorio[coluna] = sheetDados[coluna].value
                sheetRelatorio[coluna].fill = corAzul
                sheetRelatorio[coluna].border = borda
            else:
                break

        sheetRelatorio['A2'] = sheetDados['A%s' % linha].value
        sheetRelatorio['B2'] = sheetDados['B%s' % linha].value
        sheetRelatorio['C2'] = sheetDados['C%s' % linha].value

        sheetRelatorio['A2'].fill = corAmarelo
        sheetRelatorio['B2'].fill = corAmarelo
        sheetRelatorio['C2'].fill = corAmarelo

        sheetRelatorio['A2'].border = borda
        sheetRelatorio['B2'].border = borda
        sheetRelatorio['C2'].border = borda

        sheetRelatorio.column_dimensions['A'].width = 18
        sheetRelatorio.column_dimensions['B'].width = 25
        sheetRelatorio.column_dimensions['C'].width = 15

        wb_cliente.save(filename=nomeCliente+'.xlsx')
