import pandas as pd
import os
import string
import openpyxl as op

from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles import colors
from openpyxl.cell import Cell

######### Consolidando arquivo
caminhoCliente_Arq = "Clientes"
lista_Arq = list()

for arquivo in os.listdir(caminhoCliente_Arq):
    lista_Arq.append("Clientes\\"+arquivo)

dadosArquivo = pd.DataFrame()

for arquivo in lista_Arq:
    dados = pd.read_excel(arquivo)
    dadosArquivo = pd.concat([dadosArquivo, dados], ignore_index=True)

dadosArquivo.to_excel('Arquivo Consolidado.xlsx')

########## Criando Personalização

corAzul = PatternFill(start_color='0099CCFF',
                      end_color='0099CCFF',
                      fill_type='solid')

corAmarelo = PatternFill(start_color='00FFFFCC',
                      end_color='00FFFFCC',
                      fill_type='solid')

borda_fina = Side(border_style="thin",  color="000000")
borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)

########## Load para usar o op

wb_ArqConsolidado = op.load_workbook(filename='Arquivo Consolidado.xlsx')
sheetDados = wb_ArqConsolidado['Sheet1']

sheetDados.delete_cols(1)
sheetDados.title = 'Dados'

for letra in string.ascii_uppercase:
    coluna = letra + '1'

    if sheetDados[coluna].value is None:
        break
    else:
        sheetDados.column_dimensions[letra].width = 18
        sheetDados[coluna].fill = corAzul
        sheetDados[coluna].border = borda


for letra in string.ascii_uppercase:
    for tamanho in range(2, len(sheetDados[letra]) + 1):
        coluna = letra + str(tamanho)

        if sheetDados[coluna].value is None:
            break
        else:
            sheetDados[coluna].fill = corAmarelo
            sheetDados[coluna].border = borda


wb_ArqConsolidado.save(filename='Arquivo Consolidado.xlsx')

os.startfile('Arquivo Consolidado.xlsx')