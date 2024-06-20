import openpyxl as op
import os
from openpyxl.styles import PatternFill

#Criar planilha
wb = op.Workbook()
plan = wb.create_sheet('Dados')
wb.save('plan.xlsx')

#Load da mesma planilha
wb_open = op.load_workbook('plan.xlsx')
sheet = wb_open['Dados']

dados = [
    ["Nome", "Idade"],
    ["Joao", 14],
    ["Maria", 16],
    ["Luis", 19],
    ["Calopsita", 23]
]

for row in dados:
    sheet.append(row)

#Colorindo celulas
color_Title = PatternFill(start_color='00FFFF00',
                          end_color='00FFFF00',
                          fill_type='solid')

color_Value = PatternFill(start_color='0000FF00',
                          end_color='0000FF00',
                          fill_type='solid')

sheet["A1"].fill = color_Title
sheet["B1"].fill = color_Title

for linha in range(2, len(sheet['A'])+1):

    celulaColunaA = "A" + str(linha)
    celulaColunaB = "B" + str(linha)

    sheet[celulaColunaA].fill = color_Value
    sheet[celulaColunaB].fill = color_Value

wb_open.save('plan.xlsx')

os.startfile('plan.xlsx')