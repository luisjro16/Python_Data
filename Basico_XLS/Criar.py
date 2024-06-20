import openpyxl as op
import os
from openpyxl.styles import PatternFill

wb = op.Workbook()

plan = wb.create_sheet('Dados')

dados = [
    ["Nome", "Idade"],
    ["Joao", 14],
    ["Maria", 16],
    ["Luis", 19],
    ["Calopsita", 23]
]

for i, dados in enumerate(dados, start=1):
    an = "A" + str(i)
    bn = "B" + str(i)

    plan[an] = dados[0]
    plan[bn] = dados[1]


wb.save('plan.xlsx')

os.startfile('plan.xlsx')