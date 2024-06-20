from openpyxl import Workbook as WB
from openpyxl import load_workbook as load
import os

dadosSistema = load(filename='DadosSistema.xlsx')
sheetDados = dadosSistema['Dados']

wb_Relatorio = WB()
sheetRelatorio = wb_Relatorio.active

for linha in range(1, len(sheetDados['A']) +1):
    for coluna in range(1, 10):
        sheetRelatorio.cell(row=linha, column=coluna).value = sheetDados.cell(row=linha, column=coluna).value

sheetRelatorio.delete_rows(2)

sheetRelatorio.delete_cols(2)
sheetRelatorio.delete_cols(2)

sheetRelatorio.title = 'Dados Funcionarios'
wb_Relatorio.create_sheet('Resumo')

sheetResumo = wb_Relatorio['Resumo']

sheetResumo['A1'] = 'Vendedor'
sheetResumo['B1'] = 'Total vendas'

wb_Relatorio.save('Relatorio.xlsx')

os.startfile('Relatorio.xlsx')